import openpyxl, random
from thefuzz import fuzz

#fuzz.token_sort_ratio("fuzzy wuzzy was a bear", "wuzzy fuzzy was a bear")

dates = {
    1: "1-28-24",
    2: "2-4-24",
    3: "2-11-24",
    4: "2-18-24",
    5: "2-25-24",
    6: "3-3-24",
    7: "3-10-24"
}

class Student:
    def __init__(self):
        self.raw = None
        self.name = None
        self.index = None
        self.row = None
        self.relevance = None

    def __repr__(self):
        if self.relevance is None:
            return f"{self.row} - {self.name}"
        else:
            return f"{self.relevance} {self.row} - {self.name}"

    def add_session(self, number):
        slots = [r.value for r in self.raw[6:18] if r.value is None]
        choose = random.choice(slots)
        choose = dates[number]

class DiliParser():
    def __init__(self, data_path):
        self.data = openpyxl.load_workbook(data_path)
        self.dataframe = self.data.active

        self.rows = []
        self.next_row = 0

    def update(self):
        self.rows = []
        for n, r in enumerate(self.dataframe.iter_rows(0, self.dataframe.max_row)):
            name = r[1].value
            if name is None:
                continue
            student = Student()
            student.raw = r
            student.name = name
            student.index = n
            student.row = n+1
            self.rows.append(student)
        self.next_row = len(self.rows)

    def prompt(self, name):
        self.update()
        matched = []
        for student in self.rows:
            rate = fuzz.token_sort_ratio(name, student.name)
            if rate < 55:
                continue
            student.relevance = rate
            matched.append(student)
        matched = sorted(matched, reverse=True, key=lambda x: int(x.relevance))

        if len(matched) <= 0:
            pass

        elif len(matched) == 1:
            self.modify(matched[0])
            return

        for i in matched:
            if i.relevance >= 90:
                self.modify(i)
                break


    def modify(self, student):
        session = input("session: ")

        student.add_session(session)

if __name__ == "__main__":
    app = DiliParser("DILIMAN Google Sheeet.xlsx")
    n = input(">>> ")
    app.prompt(n)
